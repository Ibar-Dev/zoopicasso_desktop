import logging
import os
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from src.ventas_store import (
    archivar_ventas_activas,
    registrar_cierre,
    resumen_ventas_activas,
    ventas_activas_detalle,
)

logger = logging.getLogger(__name__)

_BASE = Path(__file__).resolve().parent.parent


def _ruta_cierres_dir() -> Path:
    valor = os.getenv("CIERRES_DIR", "").strip()
    if not valor:
        return (_BASE / "data" / "cierres").resolve()
    ruta = Path(valor).expanduser()
    if not ruta.is_absolute():
        ruta = _BASE / ruta
    return ruta.resolve()


RUTA_CIERRES = _ruta_cierres_dir()


def _generar_excel_cierre(anio_mes: str, resumen: dict) -> Path:
    RUTA_CIERRES.mkdir(parents=True, exist_ok=True)
    archivo = RUTA_CIERRES / f"cierre_mensual_{anio_mes.replace('-', '_')}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Cierre {anio_mes}"

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 18

    ws["A1"] = "Cierre mensual de ganancias"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A1:B1")

    ws["A3"] = "Periodo"
    ws["B3"] = anio_mes
    ws["A4"] = "Total activo"
    ws["B4"] = float(resumen["total"])
    ws["B4"].number_format = "#,##0.00"
    ws["A5"] = "Cantidad de ventas"
    ws["B5"] = int(resumen["cantidad_ventas"])

    ws["A7"] = "Categoría"
    ws["B7"] = "Total"
    ws["A7"].font = Font(bold=True)
    ws["B7"].font = Font(bold=True)
    ws["A7"].fill = PatternFill("solid", fgColor="D9E1F2")
    ws["B7"].fill = PatternFill("solid", fgColor="D9E1F2")

    fila = 8
    for categoria, total in sorted(resumen["por_categoria"].items()):
        ws[f"A{fila}"] = categoria
        ws[f"B{fila}"] = float(total)
        ws[f"B{fila}"].number_format = "#,##0.00"
        fila += 1

    wb.save(archivo)
    if not archivo.exists() or archivo.stat().st_size == 0:
        raise OSError(f"No se pudo verificar el Excel de cierre: {archivo}")

    return archivo


def process_monthly_closure(usuario: str) -> dict:
    anio_mes = datetime.now().strftime("%Y-%m")
    resumen = resumen_ventas_activas(anio_mes)

    if resumen["cantidad_ventas"] == 0:
        return {
            "ok": True,
            "anio_mes": anio_mes,
            "archivo_excel": None,
            "cantidad_ventas": 0,
            "total": 0.0,
            "mensaje": "No hay ventas activas para cerrar en este mes.",
        }

    detalle = ventas_activas_detalle(anio_mes)
    if not detalle:
        return {
            "ok": True,
            "anio_mes": anio_mes,
            "archivo_excel": None,
            "cantidad_ventas": 0,
            "total": 0.0,
            "mensaje": "No hay ventas activas para cerrar en este mes.",
        }

    archivo = _generar_excel_cierre(anio_mes, resumen)
    cierre_id = f"{anio_mes}-{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}"
    archived_at = datetime.now(timezone.utc).isoformat(timespec="seconds")

    actualizadas = archivar_ventas_activas(anio_mes, cierre_id, archived_at)
    registrar_cierre(
        cierre_id=cierre_id,
        anio_mes=anio_mes,
        usuario=(usuario or "").strip(),
        created_at=archived_at,
        total=float(resumen["total"]),
        cantidad_ventas=int(actualizadas),
        archivo_excel=str(archivo),
    )

    logger.info(
        "Cierre mensual completado usuario=%s periodo=%s ventas=%d total=%.2f archivo=%s",
        usuario,
        anio_mes,
        actualizadas,
        float(resumen["total"]),
        archivo,
    )

    return {
        "ok": True,
        "anio_mes": anio_mes,
        "archivo_excel": archivo.name,
        "cantidad_ventas": int(actualizadas),
        "total": float(resumen["total"]),
        "mensaje": "Cierre mensual completado correctamente.",
    }
