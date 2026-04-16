# factura_writer.py
# Genera un archivo .xlsx por factura con formato visual de factura real.
# Usa openpyxl con estilos: fuentes, bordes, rellenos y alineación.
# Los importes se consideran finales (IVA incluido).
#
# Puntos CRÍTICOS de fallo:
# - Permisos insuficientes para escribir en RUTA_FACTURAS
# - RUTA_FACTURAS_PRINCIPAL no configurada o ruta inválida
# - Copia a Windows Documents si existe, pero falla silenciosamente si no hay permisos
# - openpyxl genera excepciones si los estilos son inválidos
# - Memoria insuficiente para generar archivos grandes

import logging
import os
import shutil
import ctypes
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from src.factura_model import (
    EMAIL_EMISOR,
    DIRECCION_EMISOR,
    NOMBRE_EMISOR,
    NIF_EMISOR,
    TELEFONO_EMISOR,
    Factura,
)
from src.settings import RUTA_FACTURAS_PRINCIPAL

logger = logging.getLogger(__name__)

RUTA_FACTURAS = RUTA_FACTURAS_PRINCIPAL
logger.info(f"Módulo factura_writer inicializado. Ruta de facturas: {RUTA_FACTURAS}")
_AZUL_OSCURO = "1F4E79"   # Cabecera del emisor
_AZUL_CLARO = "BDD7EE"    # Segunda parte de la cabecera + header de tabla
_GRIS_CLARO = "F2F2F2"    # Filas alternas y bloque cliente
_BLANCO = "FFFFFF"

# ── Bordes ───────────────────────────────────────────────────────────────────
_THIN = Side(style="thin")
_BORDE = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_BORDE_TOP = Border(top=_THIN)


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _font(bold=False, size=11, color="000000") -> Font:
    return Font(bold=bold, size=size, color=color)


def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _aplicar_fila_tabla(ws, fila: int, valores: list, alternar: bool) -> None:
    """Aplica formato de fila de datos a las columnas A-D."""
    color = _GRIS_CLARO if alternar else _BLANCO
    columnas = ["A", "B", "C", "D"]
    alineaciones = ["left", "center", "right", "right"]
    for col, val, align in zip(columnas, valores, alineaciones):
        celda = ws[f"{col}{fila}"]
        celda.value = val
        celda.font = _font()
        celda.alignment = _align(h=align)
        celda.border = _BORDE
        celda.fill = _fill(color)
        if col in ("C", "D") and isinstance(val, float):
            celda.number_format = '#,##0.00'


def _rutas_documentos_windows() -> list[Path]:
    """Devuelve posibles rutas de Documentos en Windows (incluyendo fallback)."""
    if os.name != "nt":
        return []

    def _ruta_documentos_windows_sistema() -> Path | None:
        """
        Obtiene la carpeta real de Documentos usando la API de Windows.
        Esta vía funciona aunque el sistema esté en otro idioma o la ruta se haya movido.
        """
        class _GUID(ctypes.Structure):
            _fields_ = [
                ("Data1", ctypes.c_uint32),
                ("Data2", ctypes.c_uint16),
                ("Data3", ctypes.c_uint16),
                ("Data4", ctypes.c_ubyte * 8),
            ]

        folder_id = _GUID(
            0xFDD39AD0,
            0x238F,
            0x46AF,
            (ctypes.c_ubyte * 8)(0xAD, 0xB4, 0x6C, 0x85, 0x48, 0x03, 0x69, 0xC7),
        )

        try:
            path_ptr = ctypes.c_wchar_p()
            resultado = ctypes.windll.shell32.SHGetKnownFolderPath(
                ctypes.byref(folder_id),
                0,
                None,
                ctypes.byref(path_ptr),
            )
            if resultado != 0 or not path_ptr.value:
                return None

            ruta = Path(path_ptr.value)
            ctypes.windll.ole32.CoTaskMemFree(ctypes.cast(path_ptr, ctypes.c_void_p))
            return ruta
        except Exception:
            return None

    ruta_env = os.getenv("FACTURAS_DIR_WINDOWS", "").strip()
    if ruta_env:
        ruta = Path(ruta_env).expanduser()
        if not ruta.is_absolute():
            ruta = (Path.home() / ruta).resolve()
        return [ruta]

    rutas_preferidas = []
    ruta_docs_sistema = _ruta_documentos_windows_sistema()
    if ruta_docs_sistema is not None:
        rutas_preferidas.append(ruta_docs_sistema)

    base = Path.home()
    candidatas = [
        base / "Documents",
        base / "Documentos",
        base / "OneDrive" / "Documents",
        base / "OneDrive" / "Documentos",
    ]

    # OneDrive puede tener nombres como "OneDrive - Empresa"
    candidatas.extend(
        p / "Documents" for p in base.glob("OneDrive*") if p.is_dir()
    )
    candidatas.extend(
        p / "Documentos" for p in base.glob("OneDrive*") if p.is_dir()
    )

    vistas = set()
    rutas_validas = []
    for ruta in rutas_preferidas + candidatas:
        ruta_resuelta = ruta.resolve()
        if ruta_resuelta in vistas:
            continue
        vistas.add(ruta_resuelta)
        if ruta.exists() and ruta.is_dir():
            rutas_validas.append(ruta)

    # Fallback final: crear Documents bajo HOME si no se detectó ninguna carpeta.
    if not rutas_validas:
        fallback = Path.home() / "Documents"
        rutas_validas.append(fallback)
        logger.warning(
            "No se encontró carpeta de Documentos en Windows. "
            f"Se usará fallback: {fallback}"
        )

    return rutas_validas


def _copiar_en_documentos_windows(ruta_archivo: Path) -> Path | None:
    """
    Copia la factura en Documentos/Facturas en Windows, si existe Documentos.
    Falla silenciosamente si no hay permisos, pero lo logea como WARNING.
    """
    rutas_docs = _rutas_documentos_windows()
    if not rutas_docs:
        logger.debug("No se encontraron rutas de Documentos en Windows, no copiando.")
        return None

    for carpeta_documentos in rutas_docs:
        try:
            destino_dir = carpeta_documentos / "Facturas"
            destino_dir.mkdir(parents=True, exist_ok=True)
            destino = destino_dir / ruta_archivo.name
            shutil.copy2(ruta_archivo, destino)
            logger.info(f"Copia secundaria de factura guardada en: {destino}")
            return destino
        except OSError as e:
            logger.warning(
                f"No se pudo copiar factura a {carpeta_documentos}: {e}. "
                f"Verifica permisos de escritura.",
                exc_info=True
            )
        except Exception as e:
            logger.error(f"CRÍTICO: Error inesperado al copiar factura: {e}", exc_info=True)
    return None


def generar_factura_xlsx(factura: Factura) -> Path:
    """
    Genera el archivo .xlsx de la factura y lo guarda en RUTA_FACTURAS.
    Devuelve la ruta del archivo generado.
    
    Lanzará:
    - OSError si no hay permisos para escribir en RUTA_FACTURAS
    - openpyxl exceptions si hay problemas con el documento Excel
    """
    try:
        RUTA_FACTURAS.mkdir(parents=True, exist_ok=True)
    except OSError as e:
        logger.error(f"CRÍTICO: No se puede crear directorio de facturas {RUTA_FACTURAS}: {e}", exc_info=True)
        raise

    nombre_archivo = f"factura_{factura.fecha.year}_{factura.numero:03d}.xlsx"
    ruta_archivo = RUTA_FACTURAS / nombre_archivo

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Factura {factura.numero_formateado}"

    # ── Anchos de columna ─────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16

    # ── Filas 1-2: Nombre y NIF del emisor (fondo azul oscuro, texto blanco) ──
    ws.merge_cells("A1:D1")
    ws["A1"].value = NOMBRE_EMISOR
    ws["A1"].font = _font(bold=True, size=16, color=_BLANCO)
    ws["A1"].fill = _fill(_AZUL_OSCURO)
    ws["A1"].alignment = _align(h="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:D2")
    ws["A2"].value = f"NIF: {NIF_EMISOR}"
    ws["A2"].font = _font(size=10, color=_BLANCO)
    ws["A2"].fill = _fill(_AZUL_OSCURO)
    ws["A2"].alignment = _align(h="center")

    # ── Filas 3-4: Dirección y contacto (fondo azul claro) ───────────────────
    ws.merge_cells("A3:D3")
    ws["A3"].value = DIRECCION_EMISOR
    ws["A3"].font = _font(size=10)
    ws["A3"].fill = _fill(_AZUL_CLARO)
    ws["A3"].alignment = _align(h="center")

    ws.merge_cells("A4:D4")
    ws["A4"].value = f"Tel: {TELEFONO_EMISOR}   ·   {EMAIL_EMISOR}"
    ws["A4"].font = _font(size=10)
    ws["A4"].fill = _fill(_AZUL_CLARO)
    ws["A4"].alignment = _align(h="center")

    # ── Fila 5: Separador ─────────────────────────────────────────────────────
    ws.row_dimensions[5].height = 8

    # ── Fila 6: Número y fecha de factura ─────────────────────────────────────
    ws.merge_cells("A6:B6")
    ws["A6"].value = f"FACTURA  Nº  {factura.numero_formateado}"
    ws["A6"].font = _font(bold=True, size=14)
    ws["A6"].alignment = _align(h="left")
    ws.row_dimensions[6].height = 24

    ws.merge_cells("C6:D6")
    ws["C6"].value = f"Fecha: {factura.fecha_formateada}"
    ws["C6"].font = _font(size=11)
    ws["C6"].alignment = _align(h="right")

    # ── Fila 7: Separador ─────────────────────────────────────────────────────
    ws.row_dimensions[7].height = 8

    # ── Filas 8-10: Bloque cliente ────────────────────────────────────────────
    ws.merge_cells("A8:D8")
    ws["A8"].value = "DATOS DEL CLIENTE"
    ws["A8"].font = _font(bold=True, size=10, color=_BLANCO)
    ws["A8"].fill = _fill(_AZUL_OSCURO)
    ws["A8"].alignment = _align(h="left")

    ws["A9"].value = "Nombre / Empresa:"
    ws["A9"].font = _font(bold=True, size=10)
    ws["A9"].fill = _fill(_GRIS_CLARO)
    ws["A9"].alignment = _align()
    ws.merge_cells("B9:D9")
    ws["B9"].value = factura.cliente_nombre or "-"
    ws["B9"].font = _font(size=11)
    ws["B9"].fill = _fill(_GRIS_CLARO)
    ws["B9"].alignment = _align()

    ws["A10"].value = "NIF / CIF:"
    ws["A10"].font = _font(bold=True, size=10)
    ws["A10"].fill = _fill(_GRIS_CLARO)
    ws["A10"].alignment = _align()
    ws.merge_cells("B10:D10")
    ws["B10"].value = factura.cliente_nif or "-"
    ws["B10"].font = _font(size=11)
    ws["B10"].fill = _fill(_GRIS_CLARO)
    ws["B10"].alignment = _align()

    # ── Fila 11: Separador ────────────────────────────────────────────────────
    ws.row_dimensions[11].height = 8

    # ── Fila 12: Cabecera de la tabla de líneas ───────────────────────────────
    headers = ["Concepto", "Cantidad", "P. Unitario", "Total"]
    alin_headers = ["left", "center", "right", "right"]
    for col_idx, (header, ali) in enumerate(zip(headers, alin_headers), start=1):
        from openpyxl.utils import get_column_letter
        col = get_column_letter(col_idx)
        celda = ws[f"{col}12"]
        celda.value = header
        celda.font = _font(bold=True, size=10, color=_BLANCO)
        celda.fill = _fill(_AZUL_OSCURO)
        celda.alignment = _align(h=ali)
        celda.border = _BORDE
    ws.row_dimensions[12].height = 20

    # ── Filas 13+: Líneas de la factura ───────────────────────────────────────
    fila_actual = 13
    for i, linea in enumerate(factura.lineas):
        _aplicar_fila_tabla(
            ws,
            fila_actual,
            [linea.concepto, linea.cantidad, linea.precio_unitario, linea.total],
            alternar=(i % 2 == 1),
        )
        ws.row_dimensions[fila_actual].height = 18
        fila_actual += 1

    # ── Separador antes de totales ────────────────────────────────────────────
    fila_actual += 1

    # ── Totales ───────────────────────────────────────────────────────────────
    def _fila_total(fila: int, label: str, valor: float, negrita: bool = False) -> None:
        ws.merge_cells(f"A{fila}:C{fila}")
        ws[f"A{fila}"].value = label
        ws[f"A{fila}"].font = _font(bold=negrita, size=11)
        ws[f"A{fila}"].alignment = _align(h="right")
        ws[f"A{fila}"].border = _BORDE_TOP if not negrita else _BORDE
        ws[f"D{fila}"].value = valor
        ws[f"D{fila}"].font = _font(bold=negrita, size=11)
        ws[f"D{fila}"].alignment = _align(h="right")
        ws[f"D{fila}"].number_format = '#,##0.00'
        ws[f"D{fila}"].border = _BORDE if negrita else _BORDE_TOP
        ws.row_dimensions[fila].height = 18

    _fila_total(fila_actual, "Subtotal:", factura.base_imponible)
    _fila_total(fila_actual + 1, "IVA incluido:", factura.cuota_iva)
    _fila_total(fila_actual + 2, "TOTAL:", factura.total_con_iva, negrita=True)

    # Resaltar fila TOTAL
    for col in ("A", "B", "C", "D"):
        ws[f"{col}{fila_actual + 2}"].fill = _fill(_AZUL_CLARO)

    try:
        wb.save(ruta_archivo)
        logger.info(f"Factura {factura.numero_formateado} generada exitosamente en: {ruta_archivo}")
    except OSError as e:
        logger.error(f"CRÍTICO: No se pudo escribir archivo de factura en {ruta_archivo}: {e}", exc_info=True)
        raise
    except Exception as e:
        logger.error(f"CRÍTICO: Error inesperado al guardar factura {factura.numero_formateado}: {e}", exc_info=True)
        raise

    ruta_copia = _copiar_en_documentos_windows(ruta_archivo)
    if ruta_copia is not None:
        logger.info(
            f"Copia de factura {factura.numero_formateado} guardada en: {ruta_copia}"
        )

    return ruta_archivo
