from datetime import date
from pathlib import Path

import openpyxl

from src.factura_model import Factura, LineaFactura
from src import factura_writer


def _factura_demo() -> Factura:
    return Factura(
        numero=9,
        fecha=date(2026, 4, 8),
        cliente_nombre="Cliente Real",
        cliente_nif="12345678A",
        lineas=[
            LineaFactura(concepto="Servicio A", cantidad=2, precio_unitario=10.0),
            LineaFactura(concepto="Servicio B", cantidad=1, precio_unitario=5.0),
        ],
    )


def test_generar_factura_xlsx_crea_archivo(tmp_path, monkeypatch):
    monkeypatch.setattr(factura_writer, "RUTA_FACTURAS", tmp_path / "facturas")
    monkeypatch.setattr(factura_writer, "_copiar_en_documentos_windows", lambda _: None)

    ruta = factura_writer.generar_factura_xlsx(_factura_demo())

    assert ruta.exists()
    assert ruta.name == "factura_2026_009.xlsx"


def test_generar_factura_xlsx_contenido_basico(tmp_path, monkeypatch):
    monkeypatch.setattr(factura_writer, "RUTA_FACTURAS", tmp_path / "facturas")
    monkeypatch.setattr(factura_writer, "_copiar_en_documentos_windows", lambda _: None)

    ruta = factura_writer.generar_factura_xlsx(_factura_demo())

    wb = openpyxl.load_workbook(ruta)
    ws = wb.active
    assert ws["A1"].value == "Gisselle Marin Tabares"
    assert ws["A6"].value == "FACTURA  Nº  2026-009"
    assert ws["C6"].value == "Fecha: 08/04/2026"


def test_copiar_en_documentos_windows_sin_rutas(monkeypatch, tmp_path):
    archivo = tmp_path / "demo.xlsx"
    archivo.write_bytes(b"ok")
    monkeypatch.setattr(factura_writer, "_rutas_documentos_windows", lambda: [])

    resultado = factura_writer._copiar_en_documentos_windows(archivo)

    assert resultado is None


def test_copiar_en_documentos_windows_ok(monkeypatch, tmp_path):
    archivo = tmp_path / "demo.xlsx"
    archivo.write_bytes(b"ok")
    docs = tmp_path / "Documents"
    docs.mkdir(parents=True, exist_ok=True)
    monkeypatch.setattr(factura_writer, "_rutas_documentos_windows", lambda: [docs])

    resultado = factura_writer._copiar_en_documentos_windows(archivo)

    assert resultado is not None
    assert resultado.exists()
    assert resultado.parent.name == "Facturas"


def test_rutas_documentos_windows_env_relativa(monkeypatch, tmp_path):
    monkeypatch.setattr(factura_writer.os, "name", "nt", raising=False)
    monkeypatch.setenv("FACTURAS_DIR_WINDOWS", "mis-facturas")
    monkeypatch.setattr(factura_writer.Path, "home", classmethod(lambda cls: tmp_path))

    rutas = factura_writer._rutas_documentos_windows()

    assert len(rutas) == 1
    assert str(rutas[0]).endswith("mis-facturas")


def test_copia_windows_crea_fallback_documents(monkeypatch, tmp_path):
    monkeypatch.setattr(factura_writer.os, "name", "nt", raising=False)
    monkeypatch.delenv("FACTURAS_DIR_WINDOWS", raising=False)
    monkeypatch.setattr(factura_writer.Path, "home", classmethod(lambda cls: tmp_path))

    archivo = tmp_path / "factura.xlsx"
    archivo.write_bytes(b"excel")

    destino = factura_writer._copiar_en_documentos_windows(archivo)

    assert destino is not None
    assert destino.exists()
    assert (tmp_path / "Documents" / "Facturas").exists()
