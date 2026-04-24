# excel_writer.py
# Persiste cada ticket en un archivo .xlsx compatible con LibreOffice Calc.
# Una fila por línea de servicio. Si el archivo no existe, lo crea con cabeceras.

import logging
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

from tickets_src.ticket_model import Ticket

logger = logging.getLogger(__name__)

# Ruta del archivo Excel donde se acumulan todos los tickets.
_BASE = Path(__file__).resolve().parent.parent   # → raíz del proyecto
RUTA_EXCEL = _BASE / "data" / "tickets.xlsx"

# Cabeceras de la hoja. El orden debe coincidir con el de _fila_desde_linea().
CABECERAS = [
    "Nº Ticket",
    "Fecha/Hora",
    "Negocio",
    "NIF",
    "Servicio",
    "Cantidad",
    "Precio Unitario",
    "Total Línea",
    "Total Ticket",
]


def _crear_libro() -> openpyxl.Workbook:
    """
    Crea un libro nuevo con la hoja principal y las cabeceras en negrita.
    Solo se llama cuando el archivo no existe todavía.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tickets"

    ws.append(CABECERAS)

    # Negrita en la fila de cabecera para legibilidad en Calc
    for celda in ws[1]:
        celda.font = Font(bold=True)

    logger.info("Archivo Excel creado con cabeceras.")
    return wb


def _fila_desde_linea(ticket: Ticket, indice_linea: int) -> list:
    """
    Construye la lista de valores para una fila del Excel
    a partir de una línea específica del ticket.

    Args:
        ticket: Ticket completo (para acceder a cabecera y total general).
        indice_linea: Índice de la línea dentro de ticket.lineas.

    Retorna:
        list: Valores en el mismo orden que CABECERAS.
    """
    linea = ticket.lineas[indice_linea]
    return [
        ticket.numero,
        ticket.fecha_formateada,
        ticket.nombre_negocio,
        ticket.nif,
        linea.nombre,
        linea.cantidad,
        linea.precio_unitario,
        linea.total,
        ticket.total,  # Se repite en cada fila para facilitar filtros en Calc
    ]


def guardar_ticket(ticket: Ticket) -> None:
    """
    Añade todas las líneas del ticket al archivo .xlsx.
    Crea el archivo si no existe. Esta es la única función pública del módulo.

    Args:
        ticket: Ticket completo y validado listo para persistir.
    """
    RUTA_EXCEL.parent.mkdir(parents=True, exist_ok=True)

    # Carga el libro existente o crea uno nuevo
    if RUTA_EXCEL.exists():
        wb = openpyxl.load_workbook(RUTA_EXCEL)
        ws = wb.active
        logger.info(f"Añadiendo ticket #{ticket.numero} al Excel existente.")
    else:
        wb = _crear_libro()
        ws = wb.active

    # Una fila por cada línea del ticket
    for i in range(len(ticket.lineas)):
        ws.append(_fila_desde_linea(ticket, i))

    wb.save(RUTA_EXCEL)
    logger.info(f"Ticket #{ticket.numero} guardado. Líneas añadidas: {len(ticket.lineas)}.")